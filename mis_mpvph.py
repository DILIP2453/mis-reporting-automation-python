import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# =========================
# GET LATEST FILE
# =========================
def get_latest_file(pattern):

    files = glob.glob(pattern)

    if not files:
        return None

    return max(files, key=os.path.getctime)

# =========================
# FILE PATHS
# =========================
amazon_file = get_latest_file(
    r"C:\Users\Dilip\Downloads\Amazon_output_*.xlsx"
)

myntra_file = get_latest_file(
    r"C:\Users\Dilip\Downloads\Myntra_output_*.xlsx"
)

OUTPUT_FILE = r"C:\Users\Dilip\Downloads\Daily Summary Report.xlsx"

# =========================
# PRINT FILE STATUS
# =========================
if amazon_file:
    print("✅ Amazon:", amazon_file)
else:
    print("⚠️ Amazon file not found")

if myntra_file:
    print("✅ Myntra:", myntra_file)
else:
    print("⚠️ Myntra file not found")

# =========================
# DATE FILTER
# =========================
amazon_start = datetime(2026, 6, 16).date()
amazon_end = datetime(2026, 6, 16).date()

myntra_start = datetime(2026, 6, 15).date()
myntra_end = datetime(2026, 6, 16).date()
# =========================
# REPORT TIME
# =========================
report_generated_on = datetime.now().strftime("%d-%m-%Y %H:%M")

# =========================
# BORDER
# =========================
thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =========================
# EMPTY DATAFRAME TEMPLATE
# =========================
def empty_summary_df():

    return pd.DataFrame(columns=[
        "Portal",
        "Location",
        "Date",
        "Report Date",
        "Today Initiated",
        "PH Received",
        "Diff"
    ])

# =========================
# AMAZON PROCESS
# =========================
def process_amazon(file_path):

    try:

        df = pd.read_excel(
            file_path,
            sheet_name="Succesfull delivered"
        )

        required_cols = [
            "Date",
            "Location",
            "Found_In_Sheet"
        ]

        missing_cols = [
            col for col in required_cols
            if col not in df.columns
        ]

        if missing_cols:
            print(f"⚠️ Amazon skipped. Missing columns: {missing_cols}")
            return empty_summary_df()

        # DATE
        df["Date"] = pd.to_datetime(
            df["Date"],
            dayfirst=True,
            errors='coerce'
        ).dt.date

        df = df[
            (df["Date"] >= amazon_start) &
            (df["Date"] <= amazon_end)
        ]

        # IF NO DATA
        if df.empty:
            print("⚠️ Amazon file has no matching data")
            return empty_summary_df()

        # CLEAN DATA
        df["Location"] = (
            df["Location"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        df["Found_In_Sheet"] = (
            df["Found_In_Sheet"]
            .astype(str)
            .str.strip()
            .str.lower()
        )

        # SUMMARY
        summary_df = df.groupby(
            ["Location", "Date"]
        ).agg(
            Today_Initiated=("Location", "size"),
            Diff=("Found_In_Sheet", lambda x: (x == "not found").sum())
        ).reset_index()

        summary_df["PH Received"] = (
            summary_df["Today_Initiated"] - summary_df["Diff"]
        )

        summary_df["Portal"] = "Amazon"

        summary_df["Report Date"] = (
            f"{amazon_start} to {amazon_end}"
        )

        summary_df["Date"] = pd.to_datetime(
            summary_df["Date"]
        ).dt.strftime("%d-%m-%Y")

        summary_df.rename(columns={
            "Today_Initiated": "Today Initiated"
        }, inplace=True)

        return summary_df[
            [
                "Portal",
                "Location",
                "Date",
                "Report Date",
                "Today Initiated",
                "PH Received",
                "Diff"
            ]
        ]

    except Exception as e:

        print(f"⚠️ Amazon processing skipped: {e}")

        return empty_summary_df()

# =========================
# MYNTRA PROCESS
# =========================
def process_myntra(file_path):

    try:

        df = pd.read_excel(file_path)

        required_cols = [
            "deliver_to_seller_date",
            "warehouse_id",
            "Found_In_Sheet"
        ]

        missing_cols = [
            col for col in required_cols
            if col not in df.columns
        ]

        if missing_cols:
            print(f"⚠️ Myntra skipped. Missing columns: {missing_cols}")
            return empty_summary_df()

        # DATE
        df["deliver_to_seller_date"] = pd.to_datetime(
            df["deliver_to_seller_date"],
            dayfirst=True,
            errors='coerce'
        ).dt.date

        df = df[
            (df["deliver_to_seller_date"] >= myntra_start) &
            (df["deliver_to_seller_date"] <= myntra_end)
        ]

        # IF NO DATA
        if df.empty:
            print("⚠️ Myntra file has no matching data")
            return empty_summary_df()

        # CLEAN DATA
        df["warehouse_id"] = (
            df["warehouse_id"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        df["Found_In_Sheet"] = (
            df["Found_In_Sheet"]
            .astype(str)
            .str.strip()
            .str.lower()
        )

        # SUMMARY
        summary_df = df.groupby(
            ["warehouse_id", "deliver_to_seller_date"]
        ).agg(
            Today_Initiated=("warehouse_id", "size"),
            Diff=("Found_In_Sheet", lambda x: (x == "not found").sum())
        ).reset_index()

        summary_df["PH Received"] = (
            summary_df["Today_Initiated"] - summary_df["Diff"]
        )

        summary_df["Portal"] = "Myntra"

        summary_df["Report Date"] = (
            f"{myntra_start} to {myntra_end}"
        )

        summary_df["deliver_to_seller_date"] = pd.to_datetime(
            summary_df["deliver_to_seller_date"]
        ).dt.strftime("%d-%m-%Y")

        summary_df.rename(columns={
            "warehouse_id": "Location",
            "deliver_to_seller_date": "Date",
            "Today_Initiated": "Today Initiated"
        }, inplace=True)

        return summary_df[
            [
                "Portal",
                "Location",
                "Date",
                "Report Date",
                "Today Initiated",
                "PH Received",
                "Diff"
            ]
        ]

    except Exception as e:

        print(f"⚠️ Myntra processing skipped: {e}")

        return empty_summary_df()

# =========================
# PROCESS FILES
# =========================
if amazon_file:
    amazon_df = process_amazon(amazon_file)
else:
    amazon_df = empty_summary_df()

if myntra_file:
    myntra_df = process_myntra(myntra_file)
else:
    myntra_df = empty_summary_df()

# =========================
# COMBINE DATA
# =========================
summary_df = pd.concat(
    [amazon_df, myntra_df],
    ignore_index=True
)

# =========================
# GRAND TOTAL
# =========================
grand_total = {
    "Portal": "Grand Total",
    "Location": "",
    "Date": "",
    "Report Date": "",
    "Today Initiated": summary_df["Today Initiated"].sum(),
    "PH Received": summary_df["PH Received"].sum(),
    "Diff": summary_df["Diff"].sum()
}

summary_df = pd.concat(
    [summary_df, pd.DataFrame([grand_total])],
    ignore_index=True
)

# =========================
# CREATE / APPEND EXCEL
# =========================
if os.path.exists(OUTPUT_FILE):

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    start_row = ws.max_row + 3

else:

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary Report"

    start_row = 1

# =========================
# TITLE
# =========================
ws.merge_cells(
    start_row=start_row,
    start_column=1,
    end_row=start_row,
    end_column=len(summary_df.columns)
)

title_cell = ws.cell(start_row, 1)

title_cell.value = (
    f"Summary Report Generated On : {report_generated_on}"
)

title_cell.font = Font(
    size=14,
    bold=True,
    color="FFFFFF"
)

title_cell.fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

title_cell.alignment = Alignment(
    horizontal='center',
    vertical='center'
)

# =========================
# HEADERS
# =========================
for col_num, col_name in enumerate(summary_df.columns, 1):

    cell = ws.cell(start_row + 1, col_num)

    cell.value = col_name

    cell.font = Font(bold=True)

    cell.fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    cell.alignment = Alignment(
        horizontal='center',
        vertical='center'
    )

    cell.border = thin

# =========================
# DATA
# =========================
for row_num, row in enumerate(
    summary_df.values,
    start_row + 2
):

    for col_num, value in enumerate(row, 1):

        cell = ws.cell(row_num, col_num)

        cell.value = value

        cell.alignment = Alignment(
            horizontal='center',
            vertical='center'
        )

        cell.border = thin

        if str(row[0]).lower() == "grand total":

            cell.font = Font(bold=True)

            cell.fill = PatternFill(
                start_color="E2EFDA",
                end_color="E2EFDA",
                fill_type="solid"
            )

# =========================
# COLUMN WIDTH
# =========================
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 10

# =========================
# SAVE FILE
# =========================
wb.save(OUTPUT_FILE)

print("\n✅ REPORT GENERATED SUCCESSFULLY")
print("📄 File:", OUTPUT_FILE)