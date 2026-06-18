import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================
# FILE PATH
# =========================
INPUT_FILE = r"C:\Users\Dilip\Downloads\RETURN LSD 17-06-2026.xlsx"
OUTPUT_FILE = r"C:\Users\Dilip\Downloads\Daily Report.xlsx"

today = datetime.now().strftime("%d-%m-%Y")

# =========================
# LOAD DATA
# =========================
raw_df = pd.read_excel(INPUT_FILE, sheet_name="Sheet1", header=None)

# 🔍 AUTO HEADER FIND (MY REMARK BASE)
header_row = None
for i in range(len(raw_df)):
    row = raw_df.iloc[i].astype(str).str.lower()
    if row.str.contains("my remark").any():
        header_row = i
        break

if header_row is None:
    raise Exception("❌ Header not found")

df = pd.read_excel(INPUT_FILE, sheet_name="Sheet1", header=header_row)
df.columns = df.columns.astype(str).str.strip().str.lower()

# =========================
# FIND COLUMNS
# =========================
def find_col(keyword):
    for col in df.columns:
        if keyword in col:
            return col
    return None

remark_col = find_col("my remark")
location_col = find_col("location")
date_col = find_col("date")
mapping_col = find_col("mapping")

if not all([remark_col, location_col, date_col, mapping_col]):
    raise Exception("❌ Required columns missing")

# =========================
# CLEAN DATA
# =========================
df[remark_col] = df[remark_col].astype(str).str.strip().str.title()
df[mapping_col] = df[mapping_col].astype(str).str.strip().str.title()
df[location_col] = df[location_col].astype(str).str.strip().str.upper()
df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

# =========================
# DYNAMIC LOCATIONS
# =========================
locations = (
    df[location_col]
    .dropna()
    .astype(str)
    .str.strip()
    .str.upper()
)

# Remove blanks and NAN text
locations = locations[
    (locations != "") &
    (locations != "NAN")
].drop_duplicates().tolist()

print("📍 Locations Found:", locations)

# =========================
# 📄 MAPPED DONE
# =========================
mapped_df = df[df[remark_col].str.contains("mapped", case=False, na=False)]

pivot_mapped = pd.pivot_table(
    mapped_df,
    index=[date_col, mapping_col],
    columns=location_col,
    aggfunc="size",
    fill_value=0
)

pivot_mapped = pivot_mapped.sort_index(level=0)
pivot_mapped = pivot_mapped.reindex(columns=locations, fill_value=0)

pivot_mapped["Grand Total"] = pivot_mapped.sum(axis=1)

pivot_mapped = pivot_mapped.reset_index()

pivot_mapped[date_col] = pivot_mapped[date_col].dt.strftime("%d-%b")

pivot_mapped.columns = ["Returns Scanned Date","Mapping Source"] + list(pivot_mapped.columns[2:])

pivot_mapped.loc[len(pivot_mapped)] = ["Grand Total",""] + list(pivot_mapped.iloc[:,2:].sum())

# =========================
# 📄 SUMMARY (MY REMARK BASE)
# =========================
summary = pd.pivot_table(
    df,
    index=remark_col,
    columns=location_col,
    aggfunc="size",
    fill_value=0
)

summary = summary.reindex(columns=locations, fill_value=0)

summary["Grand Total"] = summary.sum(axis=1)

summary = summary.reset_index()
summary.rename(columns={remark_col:"Row Labels"}, inplace=True)

summary.loc[len(summary)] = ["Grand Total"] + list(summary.iloc[:,1:].sum())

# =========================
# 🎨 STYLE FUNCTION (FINAL FIX)
# =========================
def style_block(ws, start_row, cols, data_len):

    header_fill = PatternFill(start_color="2F75B5", fill_type="solid")
    total_fill = PatternFill(start_color="9DC3E6", fill_type="solid")

    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    align = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # HEADER
    for c in range(1, cols+1):
        cell = ws.cell(start_row+1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align
        cell.border = border

    # DATA
    for r in range(start_row+2, start_row+data_len+2):
        for c in range(1, cols+1):
            cell = ws.cell(r, c)
            cell.alignment = align
            cell.border = border

    # SPECIAL COLUMNS
    for c in range(1, cols+1):
        header = str(ws.cell(start_row+1, c).value).lower()
        if header in ["returns scanned date","mapping source","grand total","row labels"]:
            for r in range(start_row+1, start_row+data_len+2):
                cell = ws.cell(r, c)
                cell.fill = total_fill
                cell.font = bold_font

    # GRAND TOTAL ROW
    for r in range(start_row+2, start_row+data_len+2):
        if str(ws.cell(r,1).value).lower() == "grand total":
            for c in range(1, cols+1):
                cell = ws.cell(r, c)
                cell.fill = total_fill
                cell.font = bold_font

    # AUTO WIDTH
    for i in range(1, cols+1):
        max_len = 0
        col_letter = get_column_letter(i)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=i, max_col=i):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_len + 3

# =========================
# CLEAR FUNCTION
# =========================
def clear_block(ws, start_row):

    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    for row in ws.iter_rows(min_row=start_row, max_row=start_row+200, min_col=1, max_col=30):
        for cell in row:
            cell.value = None

# =========================
# WRITE FUNCTION
# =========================
def write_sheet(wb, sheet_name, data, title):

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    start_row = None
    for r in range(1, ws.max_row+1):
        if ws.cell(r,1).value == title:
            start_row = r
            break

    if start_row:
        print(f"🔄 Overwriting {sheet_name}")
        clear_block(ws, start_row)
    else:
        start_row = ws.max_row + 4
        print(f"➕ New Entry {sheet_name}")

    ws.cell(start_row,1).value = title
    ws.cell(start_row,1).font = Font(size=14, bold=True)

    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=len(data.columns))

    ws.cell(start_row,1).alignment = Alignment(horizontal="center")

    # HEADER
    for c, col in enumerate(data.columns,1):
        ws.cell(start_row+1,c).value = col

    # DATA
    for r, row in enumerate(data.values, start=start_row+2):
        for c, val in enumerate(row,1):
            ws.cell(r,c).value = val

    style_block(ws, start_row, len(data.columns), len(data))

# =========================
# SAVE
# =========================
if not os.path.exists(OUTPUT_FILE):
    wb = load_workbook(INPUT_FILE)
    wb.save(OUTPUT_FILE)

wb = load_workbook(OUTPUT_FILE)

write_sheet(wb, "Mapped Done", pivot_mapped, f"Mapped Done {today}")
write_sheet(wb, "Summary", summary, f"Data {today}")

# REMOVE DEFAULT SHEETS
for s in ["Sheet", "Sheet1", "Sheet2"]:
    if s in wb.sheetnames:
        del wb[s]

wb.save(OUTPUT_FILE)

print("✅ FINAL REPORT READY 🚀 PERFECT (NO ERROR + FULL STYLING)")